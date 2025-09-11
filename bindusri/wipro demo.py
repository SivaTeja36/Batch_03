import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font

URL = "https://careers.wipro.com/content/Early-Careers/?locale=en_US"
def scrape_wipro_early_careers():
    response = requests.get(URL)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    programs = []
    seen_titles = set()
    for header in soup.find_all("h2"):
        title = header.get_text(strip=True)

        if title in seen_titles:
            continue
        seen_titles.add(title)

        desc = []
        for sibling in header.find_next_siblings():
            if sibling.name == "h2":
                break
            if sibling.name in ["p", "div"]:
                text = sibling.get_text(strip=True)
                if text:
                    desc.append(text)

        img = header.find_previous("img")
        img_url = img["src"] if img else None

        link = header.find_next("a")
        link_url = link["href"] if link else None

        if title and desc:
            programs.append({
                "Program": title,
                "Description": " ".join(desc),
                "Image_URL": img_url,
                "Apply_Link": link_url
            })

    return programs

def save_to_excel(data, filename="wipro_early_careers.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Early Careers"

    # Headers
    headers = ["Program", "Description", "Image_URL", "Apply_Link"]
    ws.append(headers)

    for prog in data:
        ws.append([
            prog["Program"],
            prog["Description"],
            prog["Image_URL"],
            prog["Apply_Link"]
        ])

        # Make the Apply_Link clickable
        link_cell = ws.cell(row=ws.max_row, column=4)
        if prog["Apply_Link"]:
            link_cell.hyperlink = prog["Apply_Link"]
            link_cell.font = Font(color="0000FF", underline="single")  # blue clickable link

    wb.save(filename)
    print(f"✅ Scraping complete! Open '{filename}' in Excel with active links.")

if __name__ == "__main__":
    programs = scrape_wipro_early_careers()
    save_to_excel(programs)
